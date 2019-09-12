import os
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from datetime import datetime


class Handler:
    def __init__(self, qe):
        self.qe = qe
        insurance = os.path.join(r".\Excel Models", "Insurance")
        reinsurance = os.path.join(r".\Excel Models", "Reinsurance")
        capitalization = os.path.join(r".\Excel Models", "Capitalization")
        # Insurance
        if self.qe == 376:
            path = insurance + "\\376.xlsx"
        elif self.qe == 377:
            path = insurance + "\\377.xlsx"
        elif self.qe == 378:
            path = insurance + "\\378.xlsx"
        # Reinsurance
        elif self.qe == 404:
            path = reinsurance + "\\404.xlsx"
        elif self.qe == 405:
            path = reinsurance + "\\405.xlsx"
        elif self.qe == 406:
            path = reinsurance + "\\406.xlsx"
        elif self.qe == 407:
            path = reinsurance + "\\407.xlsx"
        elif self.qe == 408:
            path = reinsurance + "\\408.xlsx"
        elif self.qe == 409:
            path = reinsurance + "\\409.xlsx"
        # Capitalization
        elif self.qe == 419:
            path = capitalization + "\\419.xlsx"
        elif self.qe == 420:
            path = capitalization + "\\420.xlsx"
        elif self.qe == 421:
            path = capitalization + "\\421.xlsx"
        elif self.qe == 422:
            path = capitalization + "\\422.xlsx"
        elif self.qe == 423:
            path = capitalization + "\\423.xlsx"
        self.wb = load_workbook(os.path.abspath(path))

    def critics_to_excel(self, conn, total):
        self.report = []
        command = "SELECT "
        if self.qe == 376:
            for i in range(1, 15):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "376";'
        elif self.qe == 377:
            for i in range(1, 12):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "377";'
        elif self.qe == 378:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "378";'
        elif self.qe == 404:
            for i in range(1, 17):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "404";'
        elif self.qe == 405:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "405";'
        elif self.qe == 406:
            for i in range(1, 16):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "406";'
        elif self.qe == 407:
            for i in range(1, 12):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "407";'
        elif self.qe == 408:
            for i in range(1, 23):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "408";'
        elif self.qe == 409:
            for i in range(1, 21):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "409";'
        elif self.qe == 419:
            for i in range(1, 30):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "419";'
        elif self.qe == 410:
            for i in range(1, 8):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "420";'
        elif self.qe == 421:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "421";'
        elif self.qe == 422:
            for i in range(1, 13):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "422";'
        elif self.qe == 423:
            for i in range(1, 11):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "423";'

        for item in conn.execute(command).fetchone():
            if item is None:
                self.report.append((total,0))
            else:
                self.report.append((total,item))

    def df_to_excel(self,df,qe):
        ws = self.wb.active
        if qe == 378:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 378"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 378"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 378"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 378"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 378"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 378"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 378"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 8 - 378"]):
                ws[f"AW{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AX{row}"].value = value
                ws[f"AX{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 9 - 378"]):
                ws[f"BC{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BD{row}"].value = value
                ws[f"BD{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 10 - 378"]):
                ws[f"BI{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BI{row}"].value = value
                ws[f"BJ{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 11 - 378"]):
                ws[f"BO{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BP{row}"].value = value
                ws[f"BP{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 12 - 378"]):
                ws[f"BU{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BV{row}"].value = value
                ws[f"BV{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 13 - 378"]):
                ws[f"CA{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CB{row}"].value = value
                ws[f"CB{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 14 - 378"]):
                ws[f"CG{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CH{row}"].value = value
                ws[f"CH{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 377:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 377"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 377"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 377"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 377"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 377"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 377"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 377"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 8 - 377"]):
                ws[f"AW{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AX{row}"].value = value
                ws[f"AX{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 9 - 377"]):
                ws[f"BC{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BD{row}"].value = value
                ws[f"BD{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 10 - 377"]):
                ws[f"BI{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BI{row}"].value = value
                ws[f"BJ{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 378:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 378"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 378"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 378"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 378"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 378"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 378"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 378"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 8 - 378"]):
                ws[f"AW{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AX{row}"].value = value
                ws[f"AX{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 9 - 378"]):
                ws[f"BC{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BD{row}"].value = value
                ws[f"BD{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 10 - 378"]):
                ws[f"BI{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BI{row}"].value = value
                ws[f"BJ{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 11 - 378"]):
                ws[f"BO{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BP{row}"].value = value
                ws[f"BP{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 12 - 378"]):
                ws[f"BU{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BV{row}"].value = value
                ws[f"BV{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 13 - 378"]):
                ws[f"CA{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CB{row}"].value = value
                ws[f"CB{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 14 - 378"]):
                ws[f"CG{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CH{row}"].value = value
                ws[f"CH{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 15 - 378"]):
                ws[f"CM{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CN{row}"].value = value
                ws[f"CN{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 16 - 378"]):
                ws[f"CS{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"CT{row}"].value = value
                ws[f"CT{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 404:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 404"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 404"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 404"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 404"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 404"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 404"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 404"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 405:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 405"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 405"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 405"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 405"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 405"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 406:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 406"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 406"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 406"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 406"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 407:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 407"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 407"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 407"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 408:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 408"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 408"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 408"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 408"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 408"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 408"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 408"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 8 - 408"]):
                ws[f"AW{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AX{row}"].value = value
                ws[f"AX{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 9 - 408"]):
                ws[f"BC{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BD{row}"].value = value
                ws[f"BD{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 10 - 408"]):
                ws[f"BI{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BI{row}"].value = value
                ws[f"BJ{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 11 - 408"]):
                ws[f"BO{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BP{row}"].value = value
                ws[f"BP{row}"].number_format ="#,##0" 
                row +=1
        elif qe == 409:
            row = 14
            for index, value in zip(df.index, df["Cruzamento 1 - 409"]):
                ws[f"G{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"H{row}"].value = value
                ws[f"H{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 2 - 409"]):
                ws[f"M{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"N{row}"].value = value
                ws[f"N{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 3 - 409"]):
                ws[f"S{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"T{row}"].value = value
                ws[f"T{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 4 - 409"]):
                ws[f"Y{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"Z{row}"].value = value
                ws[f"Z{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 5 - 409"]):
                ws[f"AE{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AF{row}"].value = value
                ws[f"AF{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 6 - 409"]):
                ws[f"AK{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AK{row}"].value = value
                ws[f"AL{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 7 - 409"]):
                ws[f"AQ{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AQ{row}"].value = value
                ws[f"AR{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 8 - 409"]):
                ws[f"AW{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"AX{row}"].value = value
                ws[f"AX{row}"].number_format ="#,##0" 
                row +=1
            row = 14
            for index, value in zip(df.index, df["Cruzamento 9 - 409"]):
                ws[f"BC{row}"].value = to_excel(datetime.strptime(index,"%Y%m%d"))
                ws[f"BD{row}"].value = value
                ws[f"BD{row}"].number_format ="#,##0" 
                row +=1

    def save_xl(self,save_path): 
        self.ws = self.wb.active
        row = 14
        try:
            for line in self.report:
                self.ws[f"C{row}"].value = line[0]
                self.ws[f"D{row}"].value = line[1]
                row += 1
        except:
            pass
        save_path = os.path.abspath(
            os.path.join(
                save_path,
                self.ws.title+".xlsx"))
        self.wb.save(save_path)
    
    def qe_to_csv(self,qe_type, path, file_names):
        if qe_type == "376" or qe_type == 376:
            with open(path+"\\376_Export.csv", "a+") as csv:
                headers = [
                    "ESRSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "TPMOID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESRDATAINICIO",
                    "ESRDATAFIM",
                    "ESRDATAOCORR",
                    "ESRDATAREG",
                    "ESRVALORMOV",
                    "ESRDATACOMUNICA",
                    "ESRCODCESS",
                    "ESRNUMSIN",
                    "ESRVALORMON"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                    linha[0:7],
                                    linha[7:12],
                                    linha[12:20],
                                    linha[20:23],
                                    linha[23:27],
                                    linha[27:31],
                                    linha[31:35],
                                    linha[35:43],
                                    linha[43:51],
                                    linha[51:59],
                                    linha[59:67],
                                    linha[67:80],
                                    linha[80:88],
                                    linha[88:93],
                                    linha[93:113],
                                    linha[113:126]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "377" or qe_type == 377:
            with open(path+"\\377_Export.csv", "a+") as csv:
                headers = [
                    "ESLSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESLDATAINICIO",
                    "ESLDATAFIM",
                    "ESLDATAOCORR",
                    "ESLDATAREG",
                    "ESLVALORMOV",
                    "ESLCODCESS",
                    "ESLNUMSIN"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:20],
                                linha[20:23],
                                linha[23:27],
                                linha[27:31],
                                linha[31:39],
                                linha[39:47],
                                linha[47:55],
                                linha[55:63],
                                linha[63:76],
                                linha[76:81],
                                linha[81:101]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "378" or qe_type == 378:
            with open(path+"\\378_Export.csv", "a+") as csv:
                headers = [
                    "ESPSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "TPMOID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESPDATAINICIORO",
                    "ESPDATAFIMRO",
                    "ESPDATAEMISSRO",
                    "ESPVALORMOVRO",
                    "ESPDATAINICIORD",
                    "ESPDATAFIMRD",
                    "ESPDATAEMISSRD",
                    "ESPVALORMOVRD",
                    "ESPCODCESS",
                    "ESPFREQ",
                    "ESPVALORCARO",
                    "ESPVALORCARD",
                    "ESPVALORCIRO",
                    "ESPVALORCIRD",
                    "ESPMOEDA"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:20],
                                linha[20:23],
                                linha[23:27],
                                linha[27:31],
                                linha[31:35],
                                linha[35:43],
                                linha[43:51],
                                linha[51:59],
                                linha[59:72],
                                linha[72:80],
                                linha[80:88],
                                linha[88:96],
                                linha[96:109],
                                linha[109:114],
                                linha[114:118],
                                linha[118:131],
                                linha[131:144],
                                linha[144:157],
                                linha[157:170],
                                linha[170:173]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "404" or qe_type == 404:
            with open(path+"\\404_Export.csv", "a+") as csv:
                headers = [
                    "MSASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MSATIPOPERA",
                    "MSANUMSIN",
                    "MSANUMCONT",
                    "MSATIPOCONT",
                    "MSACODCESS",
                    "MSADATACOMUNICA",
                    "MSADATAREG",
                    "MSADARAOCORR",
                    "MSAVALORMOV",
                    "MSATIPOSIN",
                    "MSAMODCONT",
                    "MSAMOEDA",
                    "MSABASEIND",
                    "MSAVALORMON"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:21],
                                linha[21:23],
                                linha[23:24],
                                linha[24:44],
                                linha[44:70],
                                linha[70:71],
                                linha[71:76],
                                linha[76:84],
                                linha[84:92],
                                linha[92:100],
                                linha[100:113],
                                linha[113:114],
                                linha[114:116],
                                linha[116:119],
                                linha[119:120],
                                linha[120:133]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "405" or qe_type == 405:
            with open(path+"\\405_Export.csv", "a+") as csv:
                headers = [
                    "MSASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MSRNUMSIN",
                    "MSRNUMCONT",
                    "MSRTIPOCONT",
                    "MSRCODCESS",
                    "MSRDATACOMUNICA",
                    "MSRDATAREG",
                    "MSRVALOROCORR",
                    "MSRVALORMOV",
                    "MSRTIPOSIN",
                    "MSRMODCONT",
                    "MSRMOEDA",
                    "MSRDASEIND",
                    "MSRVALORMON"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:21],
                                linha[21:23],
                                linha[23:43],
                                linha[43:69],
                                linha[69:70],
                                linha[70:75],
                                linha[75:83],
                                linha[83:91],
                                linha[91:99],
                                linha[99:112],
                                linha[112:113],
                                linha[113:115],
                                linha[115:118],
                                linha[118:119],
                                linha[119:132]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "406" or qe_type == 406:
            with open(path+"\\406_Export.csv", "a+") as csv:
                headers = [
                    "SLASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "GRACODIGO",
                    "SLATIPOPERA",
                    "SLANUMSIN",
                    "SLANUMCONT",
                    "SLATIPOCONT",
                    "SLACODCESS",
                    "SLADATACOMUNICA",
                    "SLADATAREG",
                    "SLADATAOCORR",
                    "SLAVALORMOVPEN",
                    "SLAVALORMOVTOT",
                    "SLATIPOSIN",
                    "SLAMODCONT",
                    "SLAMOEDA",
                    "SLABASEIND"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:20],
                                linha[20:21],
                                linha[21:41],
                                linha[41:67],
                                linha[67:68],
                                linha[68:73],
                                linha[73:81],
                                linha[81:89],
                                linha[89:97],
                                linha[97:110],
                                linha[110:123],
                                linha[123:124],
                                linha[124:126],
                                linha[126:129],
                                linha[129:130]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "407" or qe_type == 407:
            with open(path+"\\407_Export.csv", "a+") as csv:
                headers = [
                    "SLRSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "GRACODIGO",
                    "SLRNUMSIN",
                    "SLRNUMCONT",
                    "SLRTIPOCONT",
                    "SLRCODCESS",
                    "SLRDATACOMUNICA",
                    "SLRDATAREG",
                    "SLRDATAOCORR",
                    "SLRVALORMOVPEN",
                    "SLRVALORMOVTOT",
                    "SLRTIPOSIN",
                    "SLRMODCONT",
                    "SLRMOEDA",
                    "SLRBASEIND"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:20],
                                linha[20:40],
                                linha[40:66],
                                linha[66:67],
                                linha[67:72],
                                linha[72:80],
                                linha[80:88],
                                linha[88:96],
                                linha[96:109],
                                linha[109:122],
                                linha[122:123],
                                linha[123:125],
                                linha[125:128],
                                linha[128:129]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "408" or qe_type == 408:
            with open(path+"\\408_Export.csv", "a+") as csv:
                headers = [
                    "MPRSEQ",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MPRNUMCONT",
                    "MPRNUMENDOSSO",
                    "MPRCODESS",
                    "MPRTIPOCONT",
                    "MPRMODCONT",
                    "MPRDATACEITE",
                    "MPRDATACONTR",
                    "MPRDATAINICIO",
                    "MPRDATAFIM",
                    "MPRPERCRISCO",
                    "MPRVALORMOV",
                    "MPRVALORMOVCOMIS",
                    "MPRCODCORRET",
                    "MPRVALORMOVCORRET",
                    "MPRVIGMED",
                    "MPRBASEIND",
                    "MPRMOEDA",
                    "MPRTAXACONV",
                    "MPRDATAEMISS "]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:21],
                                linha[21:23],
                                linha[23:24],
                                linha[24:50],
                                linha[50:56],
                                linha[56:61],
                                linha[61:62],
                                linha[62:64],
                                linha[64:72],
                                linha[72:80],
                                linha[80:88],
                                linha[88:96],
                                linha[96:109],
                                linha[109:115],
                                linha[115:128],
                                linha[128:141],
                                linha[141:146],
                                linha[146:148],
                                linha[148:149],
                                linha[149:152],
                                linha[152:165],
                                linha[165:173]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "409" or qe_type == 409:
            with open(path+"\\408_Export.csv", "a+") as csv:
                headers = [
                    "EMFSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "ATVCODIGO",
                    "TPFOPERADOR",
                    "FTRCODIGO",
                    "LCRCODIGO",
                    "TCTCODIGO",
                    "TPECODIGO",
                    "EMFPRAZOFLUXO",
                    "EMFVLREXPRISCO",
                    "EMFCNPJFUNDO",
                    "EMFCODISIN",
                    "EMFCODCUSTODIA",
                    "EMFMULTIPLOFATOR",
                    "EMFTXCONTRATADO",
                    "EMFTXMERCADO",
                    "TPFOPERADORDERIVATIVO",
                    "EMFVLRDERIVATIVO",
                    "EMFCODGRUPO"]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                structure = [
                                linha[0:7],
                                linha[7:12],
                                linha[12:18],
                                linha[18:21],
                                linha[21:23],
                                linha[23:49],
                                linha[49:55],
                                linha[55:60],
                                linha[60:61],
                                linha[61:63],
                                linha[63:71],
                                linha[71:79],
                                linha[79:87],
                                linha[87:95],
                                linha[95:101],
                                linha[101:114],
                                linha[114:127],
                                linha[127:132],
                                linha[132:145],
                                linha[145:147],
                                linha[147:148],
                                linha[148:151],
                                linha[151:164],
                                linha[164:172]]
                                csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass