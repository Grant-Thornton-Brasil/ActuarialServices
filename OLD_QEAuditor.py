import sqlite3
import time
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Treeview
from db import *
from QEValidations.validation import run_validations
from Web.ses import SES
from win10toast import ToastNotifier
import os
from QEMaths.maths import *
from export import qe_export
from datetime import datetime
import calendar
import threading
from glob import glob
from getpass import getuser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.datetime import to_excel


class main_window:
    def __init__(self):
        self.root = Tk()
        self.root.iconbitmap(default="icon.ico")
        self.root.title("Grant Thornton Brasil - Serviços Atuariais")
        self.root.resizable(False, False)
        self.root.config(padx=3, pady=3)
        ws = self.root.winfo_screenwidth()  # width of the screen
        hs = self.root.winfo_screenheight()  # height of the screen
        self.root.geometry(
            f"894x570+{int(round(ws/7,0))}+{int(round(hs/17,0))}")
        self.design()
        self.positions()

    def design(self):
        # Big Frame 1 - Esquerda
        self.left_frame = Frame(self.root)
        # Entcodigo
        self.entcodigo_frame = LabelFrame(
            self.left_frame, text="Código SUSEP - ENTCODIGO:"
        )
        self.entcodigo_entry = Entry(self.entcodigo_frame)
        self.entcodigo_button = Button(
            self.entcodigo_frame,
            text="Validar e obter ramos",
            command=lambda: self.get_ramos_thread(),
        )
        # Ano
        self.year_frame = LabelFrame(self.left_frame, text="Período")
        self.year_label = Label(self.year_frame, text="Ano Base:")
        self.year_spinbox = Spinbox(self.year_frame, from_=2010, to=2100)
        self.year_spinbox.delete(0, "end")
        # Ramos
        self.ramos_frame = LabelFrame(self.left_frame, text="Ramos:")
        self.ramos_text = Text(
            self.ramos_frame,
            width=23,
            height=16,
            state=DISABLED)
        self.ramos_scroll = Scrollbar(
            self.ramos_frame, command=self.ramos_text.yview)

        # Big Frame 2 - Direita
        self.right_frame = Frame(self.root)
        # QE Types
        self.qetype_frame = LabelFrame(self.right_frame, text="QE")
        self.qetype_label1 = Label(self.qetype_frame, text="Seguros:")
        self.qetype_label2 = Label(self.qetype_frame, text="Resseguros:")
        self.qetype_label3 = Label(self.qetype_frame, text="Capitalização:")
        # QE Types RBs
        self.qetype_var = IntVar()
        self.qetype_rb376 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="376", value=376
        )
        self.qetype_rb377 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="377", value=377
        )
        self.qetype_rb378 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="378", value=378
        )
        self.qetype_rb404 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="404", value=404
        )
        self.qetype_rb405 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="405", value=405
        )
        self.qetype_rb406 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="406", value=406
        )
        self.qetype_rb407 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="407", value=407
        )
        self.qetype_rb408 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="408", value=408
        )
        self.qetype_rb409 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="409", value=409
        )
        self.qetype_rb419 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="419", value=419
        )
        self.qetype_rb420 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="420", value=420
        )
        self.qetype_rb421 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="421", value=421
        )
        self.qetype_rb422 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="422", value=422
        )
        self.qetype_rb423 = Radiobutton(
            self.qetype_frame, variable=self.qetype_var, text="423", value=423
        )
        # Lista Arquivos
        self.arquivos_frame = LabelFrame(
            self.right_frame, text="Arquivos Importados")
        self.arquivos_text = Text(
            self.arquivos_frame, height=19, width=70, state=DISABLED
        )
        self.arquivos_button = Button(
            self.arquivos_frame,
            text="Importar TXTs",
            command=lambda: self.add_files(),
            height=13,
            width=12,
        )
        self.arquivos_button_clear = Button(
            self.arquivos_frame,
            text="Limpar",
            command=lambda: self.clear_files(),
            height=6,
            width=12,
        )
        self.arquivos_scroll = Scrollbar(
            self.arquivos_frame, command=self.arquivos_text.yview
        )
        # Processos
        self.processo1_var = IntVar()
        self.processo2_var = IntVar()
        self.processo3_var = IntVar()
        self.processo4_var = IntVar()
        self.processos_frame = LabelFrame(self.left_frame, text="Processos")
        self.processo1 = Checkbutton(
            self.processos_frame, text="Críticas", variable=self.processo1_var
        )
        self.processo2 = Checkbutton(
            self.processos_frame,
            text="Confronto ($)",
            variable=self.processo2_var)
        self.processo3 = Checkbutton(
            self.processos_frame,
            text="Críticas (Detalhamento)",
            variable=self.processo3_var,
            command=lambda: self.processo1_var.set(1),
        )
        self.processo4 = Checkbutton(
            self.processos_frame,
            text="Exportar para CSV",
            variable=self.processo4_var)

        # Path Bars:
        self.path_bars_frame = LabelFrame(self.right_frame, text="Caminhos")
        # Confrontros ($)
        self.path_confrontos_label = Label(
            self.path_bars_frame, text="Confrontos ($)")
        self.path_confrontos_entry = Entry(
            self.path_bars_frame,
            width=73,
            state=DISABLED,
            disabledbackground="white",
            disabledforeground="black",
        )
        self.path_confrontos_button = Button(
            self.path_bars_frame,
            text="Selecionar Pasta",
            command=lambda: self.get_folders(0),
        )
        # Criticas (Detalhamento)
        self.path_detalhamento_label = Label(
            self.path_bars_frame, text="Criticas (Detalhamento)"
        )
        self.path_detalhamento_entry = Entry(
            self.path_bars_frame,
            width=73,
            state=DISABLED,
            disabledbackground="white",
            disabledforeground="black",
        )
        self.path_detalhamento_button = Button(
            self.path_bars_frame,
            text="Selecionar Pasta",
            command=lambda: self.get_folders(1),
        )
        # Exportação (CSVs)
        self.path_export_label = Label(
            self.path_bars_frame, text="Exportação CSVs")
        self.path_export_entry = Entry(
            self.path_bars_frame,
            width=73,
            state=DISABLED,
            disabledbackground="white",
            disabledforeground="black",
        )
        self.path_export_button = Button(
            self.path_bars_frame,
            text="Selecionar Pasta",
            command=lambda: self.get_folders(2),
        )

        # Tree Process
        self.process_label = Label(self.root, text="Status:")
        self.process_tree_frame = Frame(self.root)
        self.process_start = Button(
            self.process_tree_frame,
            height=2,
            text="EXECUTAR!",
            command=lambda: self.validate(),
            font=("TkDefaultFont", 9, "bold"),
        )

    def positions(self):
        # Big Frame 1 - Esquerda
        self.left_frame.grid(row=0, column=0, sticky=N)
        # Entcodigo
        self.entcodigo_frame.pack(fill=BOTH)
        self.entcodigo_entry.pack(fill=BOTH)
        self.entcodigo_button.pack(fill=BOTH)
        # Ano
        self.year_frame.pack(fill=X, ipady=3)
        self.year_label.grid(row=0, column=0)
        self.year_spinbox.grid(row=0, column=1)

        # Ramos
        self.ramos_frame.pack(fill=BOTH)
        self.ramos_text.pack(side=LEFT)
        self.ramos_scroll.pack(side=LEFT, fill=Y)
        self.ramos_text.config(yscrollcommand=self.ramos_scroll.set)

        # Big Frame 2 - Direita
        self.right_frame.grid(row=0, column=1)
        self.qetype_frame.pack(fill=X)
        self.qetype_label1.grid(row=0, column=0, sticky=W)
        self.qetype_label2.grid(row=1, column=0, sticky=W)
        self.qetype_label3.grid(row=2, column=0, sticky=W)
        self.qetype_rb376.grid(row=0, column=1)
        self.qetype_rb377.grid(row=0, column=2)
        self.qetype_rb378.grid(row=0, column=3)
        self.qetype_rb404.grid(row=1, column=1)
        self.qetype_rb405.grid(row=1, column=2)
        self.qetype_rb406.grid(row=1, column=3)
        self.qetype_rb407.grid(row=1, column=4)
        self.qetype_rb408.grid(row=1, column=5)
        self.qetype_rb409.grid(row=1, column=6)
        self.qetype_rb419.grid(row=2, column=1)
        self.qetype_rb420.grid(row=2, column=2)
        self.qetype_rb421.grid(row=2, column=3)
        self.qetype_rb422.grid(row=2, column=4)
        self.qetype_rb423.grid(row=2, column=5)
        # Arquivos
        self.arquivos_frame.pack(fill=X)
        self.arquivos_text.grid(row=0, column=0)
        self.arquivos_scroll.grid(row=0, column=1, sticky=(N, W, E, S))
        self.arquivos_button.grid(row=0, column=2, sticky=N, columnspan=2)
        self.arquivos_button_clear.grid(
            row=0, column=2, sticky=S, columnspan=2)
        self.arquivos_text.config(yscrollcommand=self.arquivos_scroll.set)
        # Processos
        self.processos_frame.pack(fill=BOTH, side=BOTTOM, anchor=S)
        self.processo1.grid(row=0, column=0, sticky=W, pady=2)
        self.processo2.grid(row=1, column=0, sticky=W, pady=2)
        self.processo3.grid(row=2, column=0, sticky=W, pady=2)
        self.processo4.grid(row=3, column=0, sticky=W, pady=2)

        # Path Bars
        self.path_bars_frame.pack(fill=X)
        self.path_confrontos_label.grid(row=0, column=0, sticky=W, pady=1)
        self.path_confrontos_entry.grid(row=0, column=1, pady=1)
        self.path_confrontos_button.grid(row=0, column=2, padx=3, pady=1)

        self.path_detalhamento_label.grid(row=1, column=0, sticky=W, pady=1)
        self.path_detalhamento_entry.grid(row=1, column=1, pady=1)
        self.path_detalhamento_button.grid(row=1, column=2, padx=3, pady=1)
        self.path_export_label.grid(row=2, column=0, sticky=W, pady=1)
        self.path_export_entry.grid(row=2, column=1, pady=1)
        self.path_export_button.grid(row=2, column=2, pady=1)

        # Tree Process
        self.process_tree_frame.grid(
            row=1, column=0, columnspan=2, sticky=(W, E))
        self.process_start.pack(fill=BOTH)

    # COMMANDS

    def get_folders(self, process_type):
        if process_type == 0:
            self.path_confrontos_entry.config(state=NORMAL)
            self.path_confrontos_entry.delete(0, END)
            self.path_confrontos_entry.insert(
                0, string=filedialog.askdirectory())
            self.path_confrontos_entry.config(state=DISABLED)
        if process_type == 1:
            self.path_detalhamento_entry.config(state=NORMAL)
            self.path_detalhamento_entry.delete(0, END)
            self.path_detalhamento_entry.insert(
                0, string=filedialog.askdirectory())
            self.path_detalhamento_entry.config(state=DISABLED)
        elif process_type == 2:
            self.path_export_entry.config(state=NORMAL)
            self.path_export_entry.delete(0, END)
            self.path_export_entry.insert(0, string=filedialog.askdirectory())
            self.path_export_entry.config(state=DISABLED)

    def get_ramos_thread(self):
        x = threading.Thread(target=self.validate_entcodigo)
        x.start()

    def validate_entcodigo(self):
        try:
            int(self.entcodigo_entry.get())
            if len(self.entcodigo_entry.get()) != 5:
                raise
            int(self.year_spinbox.get())
        except BaseException:
            messagebox.showerror(
                title="Erro",
                message="Verifique o ENTCODIGO e o ANO!\n\n"
                "ENTCODIGO -> (Inteiro e cinco digitos)",
            )
            return
        self.ramos_text.delete("1.0", END)
        self.entcodigo_button.config(text="Acessando SES...", state=DISABLED)
        ramos = SES()
        ramos = ramos.get_ramos(
            self.entcodigo_entry.get(),
            self.year_spinbox.get())
        if not ramos:
            messagebox.showerror(
                title="Erro",
                message="Impossível validar o ENTCODIGO"
                " e consultar os ramos.\n\n"
                "Verifique sua conexão...\n\n"
                "OBS: É possível imputar manualmente.",
            )
            self.ramos_text.config(state=NORMAL)
            self.entcodigo_button.config(
                text="Validar e obter ramos", state=NORMAL)
            return
        self.ramos_text.config(state=NORMAL)
        for ramo in ramos:
            self.ramos_text.insert(END, ramo + "\n")
        self.ramos_text.config(state=DISABLED)
        self.entcodigo_button.config(
            text="Validar e obter ramos", state=NORMAL)

    def validate(self):
        # ENTCODIGO
        if (
            self.entcodigo_entry.get() is None
            or self.entcodigo_entry.get() == ""
            or len(self.entcodigo_entry.get()) == 0
        ):
            messagebox.showerror(title="Erro", message="Insira um ENTCODIGO.")
            return
        # Year
        if (
            self.year_spinbox.get() == 0
            or self.year_spinbox.get() is None
            or self.year_spinbox.get() == ""
        ):
            messagebox.showerror(title="Erro", message="Insira um Ano Base.")
            return
        # Ramos
        try:
            if len(self.ramos_text.get("1.0", END).split("\n")[:-1]) == 1:
                raise
        except BaseException:
            messagebox.showerror(
                title="Erro", message="Preencha os ramos\n\n"
                "É possível preencher manualmente.", )
            return
        # QE
        if self.qetype_var.get() == 0:
            messagebox.showerror(
                title="Erro", message="Selecione um Quadro Estatístico"
            )
            return
        # Processos
        if (
            self.processo1_var.get() == 0
            and self.processo2_var.get() == 0
            and self.processo3_var.get() == 0
            and self.processo4_var.get() == 0
        ):
            messagebox.showerror(
                title="Erro", message="Selecione aom menos um processo."
            )
            return
        args = (
            getuser().split(".")[0].title()
            + " - "
            + datetime.now().strftime("%d%b%y-%Hh%Mm%Ss%f")
        )
        thread = threading.Thread(target=self.run, args=[args])
        thread.start()

    def add_files(self):
        arquivos = filedialog.askopenfilenames()
        self.arquivos_text.config(state=NORMAL)
        for n, arquivo in enumerate(arquivos, 1):
            self.arquivos_text.insert(END, str(n) + "\n")
            self.arquivos_text.insert(END, arquivo + "\n")
            self.arquivos_text.insert(END, "-" * 70 + "\n")
        self.arquivos_text.config(state=DISABLED)

    def clear_files(self):
        self.arquivos_text.config(state=NORMAL)
        self.arquivos_text.delete("1.0", END)
        self.arquivos_text.config(state=DISABLED)

    def run(self, process_number):
        self.process_start.config(state=DISABLED, text="PROCESSANDO...")
        conn = sqlite3.connect(f"DBs\\{process_number}.db")
        start = time.time()
        processos = [
            self.processo1_var.get(),
            self.processo2_var.get(),
            self.processo3_var.get(),
            self.processo4_var.get(),
        ]
        # Críticas
        if processos[0] == 1:
            qe = self.qetype_var.get()
            entcodigo = self.entcodigo_entry.get()
            ramcodigos = self.ramos_text.get("1.0", END).splitlines()[:-1]
            create_main_tables(conn)
            esrcodcess = [
                "38741",
                "30074",
                "34819",
                "36099",
                "37052",
                "38253",
                "31623",
                "38873",
                "33294",
                "39764",
                "37729",
                "31551",
                "38270",
                "30201",
                "34665",
                "32875",
                entcodigo,
            ]
            for arquivo in self.arquivos_text.get(
                    "1.0", END).splitlines()[:-1]:
                if arquivo != "-" * 70:
                    try:
                        with open(arquivo, "r") as txt:

                            linhas = txt.readlines()
                            for n, linha in enumerate(linhas, 1):
                                linha = linha.replace(",", ".").strip()
                                run_validations(
                                    qe=qe,
                                    nome_arquivo=txt.name,
                                    linha=linha,
                                    n=n,
                                    conn=conn,
                                    year=int(self.year_spinbox.get()),
                                    entcodigo=entcodigo,
                                    ramcodigos=ramcodigos,
                                    esrcodcess=esrcodcess,
                                )
                    except FileNotFoundError:
                        pass
            conn.commit()
        # Confrontos
        if processos[1] == 1:
            year = int(self.year_spinbox.get())
            dates_seguros = [
                datetime(year, month, calendar.monthrange(year, month)[1]).strftime(
                    "%Y%m%d"
                )
                for month in range(1, 13)
            ]
            dates_reseguros = [
                f"2018" +
                f"{month}".zfill(2) for month in range(
                    1,
                    13)]
            m376 = maths_376(dates_seguros)
            m377 = maths_377(dates_seguros)
            m378 = maths_378(dates_seguros)
            m404 = maths_404(dates_reseguros)
            m405 = maths_405(dates_reseguros)
            m406 = maths_406(dates_reseguros)
            m407 = maths_407(dates_reseguros)
            m408 = maths_408(dates_reseguros)
            m409 = maths_408(dates_reseguros)
            qe = self.qetype_var.get()
            for arquivo in self.arquivos_text.get(
                    "1.0", END).splitlines()[:-1]:
                try:
                    with open(arquivo, "r") as txt:
                        linhas = txt.readlines()
                        for linha in linhas:
                            linha = linha.replace(",", ".").strip()
                            if qe == 376:
                                m376.run(linha)
                            elif qe == 377:
                                m377.run(linha)
                            elif qe == 376:
                                m378.run(linha)
                            elif qe == 404:
                                m404.run(linha)
                            elif qe == 405:
                                m405.run(linha)
                            elif qe == 406:
                                m406.run(linha)
                            elif qe == 407:
                                m407.run(linha)
                            elif qe == 408:
                                m408.run(linha)
                            elif qe == 409:
                                m409.run(linha)
                except FileNotFoundError:
                    pass
                except KeyError:
                    messagebox.showerror(
                        title="Erro",
                        message="As datas dos campos MRFMESANO não condizem "
                        "com o ano base informado.\n\n"
                        "Favor corrigir o ano base ou selecione"
                        "novamente os arquivos!",
                    )
            if qe == 376 or qe == 377 or qe == 378:
                wb_insurance = load_workbook(
                    os.path.abspath(
                        os.path.join(
                            "Excel Models",
                            "Modelo Seguros 2019.xlsx")))
                ws376 = wb_insurance["Validação 376"]
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 1 - 376"]):
                    ws376["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 2 - 376"]):
                    ws376["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 3 - 376"]):
                    ws376["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 4 - 376"]):
                    ws376["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 5 - 376"]):
                    ws376["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 6 - 376"]):
                    ws376["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 7 - 376"]):
                    ws376["AQ" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["AR" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 8 - 376"]):
                    ws376["AW" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["AY" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 9 - 376"]):
                    ws376["BC" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["BD" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 10 - 376"]):
                    ws376["BI" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["BJ" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 11 - 376"]):
                    ws376["BO" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["BP" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 12 - 376"]):
                    ws376["BU" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["BV" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 13 - 376"]):
                    ws376["CA" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["CB" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m376.df.index, m376.df["Cruzamento 14 - 376"]):
                    ws376["CG" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws376["CH" + str(rows)].value = df_row
                    rows += 1

                ws377 = wb_insurance["Validação 377"]
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 1 - 377"]):
                    ws377["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 2 - 377"]):
                    ws377["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 3 - 377"]):
                    ws377["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 4 - 377"]):
                    ws377["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 5 - 377"]):
                    ws377["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 6 - 377"]):
                    ws377["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 7 - 377"]):
                    ws377["AQ" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["AR" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 8 - 377"]):
                    ws377["AW" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["AY" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 9 - 377"]):
                    ws377["BC" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["BD" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m377.df.index, m377.df["Cruzamento 10 - 377"]):
                    ws377["BI" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws377["BJ" + str(rows)].value = df_row
                    rows += 1
                rows = 14

                ws378 = wb_insurance["Validação 378"]
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 1 - 378"]):
                    ws378["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 2 - 378"]):
                    ws378["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 3 - 378"]):
                    ws378["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 4 - 378"]):
                    ws378["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 5 - 378"]):
                    ws378["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 6 - 378"]):
                    ws378["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 7 - 378"]):
                    ws378["AQ" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["AR" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 8 - 378"]):
                    ws378["AW" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["AY" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 9 - 378"]):
                    ws378["BC" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["BD" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 10 - 378"]):
                    ws378["BI" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["BJ" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 11 - 378"]):
                    ws378["BO" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["BP" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 12 - 378"]):
                    ws378["BU" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["BV" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 13 - 378"]):
                    ws378["CA" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["CB" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 14 - 378"]):
                    ws378["CG" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["CH" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 15 - 378"]):
                    ws378["CM" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["CN" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m378.df.index, m378.df["Cruzamento 16 - 378"]):
                    ws378["CS" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws378["CT" + str(rows)].value = df_row
                    rows += 1

                wb_insurance.save(
                    os.path.abspath(
                        os.path.join(
                            self.path_confrontos_entry.get(), "Insurances.xlsx"
                        )
                    )
                )
            elif qe in [404, 405, 406, 407, 408, 409]:
                wb_insurance = load_workbook(
                    os.path.abspath(
                        os.path.join(
                            "Excel Models",
                            "Modelo Resseguros 2019.xlsx")))
                ws404 = wb_insurance["Validação 404"]
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 1 - 404"]):
                    ws404["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 2 - 404"]):
                    ws404["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 3 - 404"]):
                    ws404["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 4 - 404"]):
                    ws404["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 5 - 404"]):
                    ws404["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 6 - 404"]):
                    ws404["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m404.df.index, m404.df["Cruzamento 7 - 404"]):
                    ws404["AQ" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws404["AR" + str(rows)].value = df_row
                    rows += 1
                ws405 = wb_insurance["Validação 405"]
                rows = 14
                for index, df_row in zip(
                        m405.df.index, m405.df["Cruzamento 1 - 405"]):
                    ws405["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws405["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m405.df.index, m405.df["Cruzamento 2 - 405"]):
                    ws405["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws405["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m405.df.index, m405.df["Cruzamento 3 - 405"]):
                    ws405["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws405["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m405.df.index, m405.df["Cruzamento 4 - 405"]):
                    ws405["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws405["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m405.df.index, m405.df["Cruzamento 5 - 405"]):
                    ws405["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws405["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                ws406 = wb_insurance["Validação 406"]
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 1 - 406"]):
                    ws406["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 2 - 406"]):
                    ws406["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 3 - 406"]):
                    ws406["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 4 - 406"]):
                    ws406["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 5 - 406"]):
                    ws406["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m406.df.index, m406.df["Cruzamento 6 - 406"]):
                    ws406["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws406["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                ws407 = wb_insurance["Validação 407"]
                rows = 14
                for index, df_row in zip(
                        m407.df.index, m407.df["Cruzamento 1 - 407"]):
                    ws407["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws407["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m407.df.index, m407.df["Cruzamento 2 - 407"]):
                    ws407["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws407["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m407.df.index, m407.df["Cruzamento 3 - 407"]):
                    ws407["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws407["T" + str(rows)].value = df_row
                    rows += 1
                ws408 = wb_insurance["Validação 408"]
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 1 - 408"]):
                    ws408["G" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["H" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 2 - 408"]):
                    ws408["M" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["N" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 3 - 408"]):
                    ws408["S" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["T" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 4 - 408"]):
                    ws408["Y" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["Z" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 5 - 408"]):
                    ws408["AE" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["AF" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 6 - 408"]):
                    ws408["AK" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["AL" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 7 - 408"]):
                    ws408["AQ" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["AR" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 8 - 408"]):
                    ws408["AW" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["AX" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 9 - 408"]):
                    ws408["BC" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["BD" + str(rows)].value = df_row
                    rows += 1
                rows = 14
                for index, df_row in zip(
                        m408.df.index, m408.df["Cruzamento 10 - 408"]):
                    ws408["BI" + str(rows)].value = to_excel(
                        datetime.strptime(index, "%Y%m%d")
                    )
                    ws408["BJ" + str(rows)].value = df_row
                    rows += 1
                wb_insurance.save(
                    os.path.abspath(
                        os.path.join(
                            self.path_confrontos_entry.get(),
                            "Reinsurances.xlsx")))
            elif qe in [419, 420, 421, 422, 423]:
                pass

        # Detalhamento
        if processos[2] == 1:
            path = os.path.abspath(filedialog.askdirectory())
            make_report(self.qetype_var.get(), conn, path)
        # export
        if processos[3] == 1:
            path = os.path.abspath(filedialog.askdirectory())
            qe_export(
                self.qetype_var.get(),
                path,
                self.arquivos_text.get("1.0", END).splitlines(),
            )
        end = time.time()
        self.process_start.config(state=NORMAL, text="EXECUTAR!")
        notifier = ToastNotifier()
        notifier.show_toast(
            "Processo Finalizado com Sucesso!",
            f"O processo levou {round(end-start,2)} segundos",
            icon_path="icon.ico",
            threaded=True,
        )


if __name__ == "__main__":
    for db in glob("DBs\\*.db"):
        os.remove(db)
    a = main_window()
    a.root.mainloop()