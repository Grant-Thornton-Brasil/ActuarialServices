import os
from numpy import array
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from datetime import datetime
import pyodbc


class Handler:
    def __init__(self, qe):
        self.qe = qe
        insurance = os.path.join("Excel Models", "Insurance")
        reinsurance = os.path.join("Excel Models", "Reinsurance")
        capitalization = os.path.join("Excel Models", "Capitalization")
        # Insurance
        if self.qe == 376:
            path = insurance + "\\376.xlsx"
        elif self.qe == 377:
            path = insurance + "\\377.xlsx"
        elif self.qe == 378:
            path = insurance + "\\378.xlsx"
        # Reinsurance
        elif self.qe == 404:
            path = reinsurance + "\\404.xlsx"
        elif self.qe == 405:
            path = reinsurance + "\\405.xlsx"
        elif self.qe == 406:
            path = reinsurance + "\\406.xlsx"
        elif self.qe == 407:
            path = reinsurance + "\\407.xlsx"
        elif self.qe == 408:
            path = reinsurance + "\\408.xlsx"
        elif self.qe == 409:
            path = reinsurance + "\\409.xlsx"
        # Capitalization
        elif self.qe == 419:
            path = capitalization + "\\419.xlsx"
        elif self.qe == 420:
            path = capitalization + "\\420.xlsx"
        elif self.qe == 421:
            path = capitalization + "\\421.xlsx"
        elif self.qe == 422:
            path = capitalization + "\\422.xlsx"
        elif self.qe == 423:
            path = capitalization + "\\423.xlsx"
        self.wb = load_workbook(os.path.abspath(path))

    def get_from_fip(self, db_path, year, cmpid, entcodigo):
        values = []
        self.conn = pyodbc.connect(
            r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + db_path
        )
        query = self.conn.execute(
            f"""SELECT sum(vdrValor)
                FROM ValoresMovRamos
                WHERE ((mrfMesAno >= #01/01/{year}#)
                AND (mrfMesAno <= #31/12/{year}#)
                AND (CMPID = {cmpid})
                AND (entCodigo = '{entcodigo}'))
                GROUP BY mrfMesAno
                ORDER BY mrfMesAno
                """
        ).fetchall()
        if len(query) == 0:
            query = self.conn.execute(
                f"""SELECT sum(vdrvalor)
                    FROM ValoresResMovGrupos
                    WHERE ((mrfmesano >= #01/01/{year}#)
                    AND (mrfmesano <= #31/12/{year}#)
                    AND (cmpid = {cmpid})
                    AND (entcodigo = '{entcodigo}'))
                    GROUP BY mrfmesano
                    ORDER BY mrfmesano
                    """
            ).fetchall()
        for i in query:
            values.append(i[0])
        return array(values)

    def critics_to_excel(self, conn, total):
        self.report = []
        command = "SELECT "
        if self.qe == 376:
            for i in range(1, 15):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "376";'
        elif self.qe == 377:
            for i in range(1, 12):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "377";'
        elif self.qe == 378:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "378";'
        elif self.qe == 404:
            for i in range(1, 17):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "404";'
        elif self.qe == 405:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "405";'
        elif self.qe == 406:
            for i in range(1, 16):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "406";'
        elif self.qe == 407:
            for i in range(1, 12):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "407";'
        elif self.qe == 408:
            for i in range(1, 23):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "408";'
        elif self.qe == 409:
            for i in range(1, 21):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "409";'
        elif self.qe == 419:
            for i in range(1, 30):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "419";'
        elif self.qe == 420:
            for i in range(1, 16):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "420";'
        elif self.qe == 421:
            for i in range(1, 14):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "421";'
        elif self.qe == 422:
            for i in range(1, 13):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "422";'
        elif self.qe == 423:
            for i in range(1, 11):
                command += f"SUM(T{i}), "
            command = command.strip()[:-1] + ' FROM "423";'

        for item in conn.execute(command).fetchone():
            if item is None:
                self.report.append((total, 0))
            else:
                self.report.append((total, item))

    def df_to_excel(self, df_cruz, qe, db_path, year, entcodigo):
        ws = self.wb.active
        if qe == 376:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 376"],
                self.get_from_fip(db_path, year, 12184, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 376"],
                self.get_from_fip(db_path, year, 12185, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 376"],
                self.get_from_fip(db_path, year, 12186, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 376"],
                self.get_from_fip(db_path, year, 12188, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 376"],
                self.get_from_fip(db_path, year, 12189, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 6 - 376"],
                self.get_from_fip(db_path, year, 12190, entcodigo),
            ):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 376"],
                self.get_from_fip(db_path, year, 12218, entcodigo),
            ):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 8 - 376"],
                self.get_from_fip(db_path, year, 12219, entcodigo),
            ):
                ws[f"AW{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AW{row}"].offset(column=1).value = value
                ws[f"AW{row}"].offset(column=2).value = value_fip
                ws[f"AW{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 9 - 376"],
                self.get_from_fip(db_path, year, 12208, entcodigo),
            ):
                ws[f"BC{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BC{row}"].offset(column=1).value = value
                ws[f"BC{row}"].offset(column=2).value = value_fip
                ws[f"BC{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 10 - 376"],
                self.get_from_fip(db_path, year, 12209, entcodigo),
            ):
                ws[f"BI{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BI{row}"].offset(column=1).value = value
                ws[f"BI{row}"].offset(column=2).value = value_fip
                ws[f"BI{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 11 - 376"],
                self.get_from_fip(db_path, year, 12224, entcodigo),
            ):
                ws[f"BO{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BO{row}"].offset(column=1).value = value
                ws[f"BO{row}"].offset(column=2).value = value_fip
                ws[f"BO{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 12 - 376"],
                self.get_from_fip(db_path, year, 12225, entcodigo),
            ):
                ws[f"BU{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BU{row}"].offset(column=1).value = value
                ws[f"BU{row}"].offset(column=2).value = value_fip
                ws[f"BU{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BU{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BU{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 13 - 376"],
                self.get_from_fip(db_path, year, 12234, entcodigo),
            ):
                ws[f"CA{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CA{row}"].offset(column=1).value = value
                ws[f"CA{row}"].offset(column=2).value = value_fip
                ws[f"CA{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CA{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CA{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 14 - 376"],
                self.get_from_fip(db_path, year, 12235, entcodigo),
            ):
                ws[f"CG{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CG{row}"].offset(column=1).value = value
                ws[f"CG{row}"].offset(column=2).value = value_fip
                ws[f"CG{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CG{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CG{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 377:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 377"],
                self.get_from_fip(db_path, year, 12278, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 377"],
                self.get_from_fip(db_path, year, 12279, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 377"],
                self.get_from_fip(db_path, year, 12280, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 377"],
                self.get_from_fip(db_path, year, 12300, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 377"],
                self.get_from_fip(db_path, year, 12282, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 6 - 377"],
                self.get_from_fip(db_path, year, 12283, entcodigo),
            ):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 377"],
                self.get_from_fip(db_path, year, 12284, entcodigo),
            ):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 8 - 377"],
                self.get_from_fip(db_path, year, 12301, entcodigo),
            ):
                ws[f"AW{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AW{row}"].offset(column=1).value = value
                ws[f"AW{row}"].offset(column=2).value = value_fip
                ws[f"AW{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 9 - 377"],
                self.get_from_fip(db_path, year, 12326, entcodigo),
            ):
                ws[f"BC{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BC{row}"].offset(column=1).value = value
                ws[f"BC{row}"].offset(column=2).value = value_fip
                ws[f"BC{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 10 - 377"],
                self.get_from_fip(db_path, year, 12332, entcodigo),
            ):
                ws[f"BI{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BI{row}"].offset(column=1).value = value
                ws[f"BI{row}"].offset(column=2).value = value_fip
                ws[f"BI{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 378:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 378"],
                self.get_from_fip(db_path, year, 12012, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 378"],
                self.get_from_fip(db_path, year, 12016, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 378"],
                self.get_from_fip(db_path, year, 12020, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 378"],
                self.get_from_fip(db_path, year, 12034, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 378"],
                self.get_from_fip(db_path, year, 12014, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index, 
                df_cruz["Cruzamento 6 - 378"], 
                self.get_from_fip(db_path, year, 12018, entcodigo)):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 378"],
                self.get_from_fip(db_path, year, 12022, entcodigo)):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 8 - 378"],
                self.get_from_fip(db_path, year, 12036, entcodigo),
            ):
                ws[f"AW{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"AW{row}"].offset(column=1).value = value
                ws[f"AW{row}"].offset(column=2).value = value_fip
                ws[f"AW{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 9 - 378"],
                self.get_from_fip(db_path, year, 12013, entcodigo),
            ):
                ws[f"BC{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BC{row}"].offset(column=1).value = value
                ws[f"BC{row}"].offset(column=2).value = value_fip
                ws[f"BC{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 10 - 378"],
                self.get_from_fip(db_path, year, 12017, entcodigo),
            ):
                ws[f"BI{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BI{row}"].offset(column=1).value = value
                ws[f"BI{row}"].offset(column=2).value = value_fip
                ws[f"BI{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 11 - 378"],
                self.get_from_fip(db_path, year, 12021, entcodigo),
            ):
                ws[f"BO{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BO{row}"].offset(column=1).value = value
                ws[f"BO{row}"].offset(column=2).value = value_fip
                ws[f"BO{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 12 - 378"],
                self.get_from_fip(db_path, year, 12035, entcodigo),
            ):
                ws[f"BU{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"BU{row}"].offset(column=1).value = value
                ws[f"BU{row}"].offset(column=2).value = value_fip
                ws[f"BU{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BU{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BU{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 13 - 378"],
                self.get_from_fip(db_path, year, 12026, entcodigo),
            ):
                ws[f"CA{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CA{row}"].offset(column=1).value = value
                ws[f"CA{row}"].offset(column=2).value = value_fip
                ws[f"CA{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CA{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CA{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 14 - 378"],
                self.get_from_fip(db_path, year, 12481, entcodigo),
            ):
                ws[f"CG{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CG{row}"].offset(column=1).value = value
                ws[f"CG{row}"].offset(column=2).value = value_fip
                ws[f"CG{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CG{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CG{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 15 - 378"],
                self.get_from_fip(db_path, year, 12482, entcodigo),
            ):
                ws[f"CM{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CM{row}"].offset(column=1).value = value
                ws[f"CM{row}"].offset(column=2).value = value_fip
                ws[f"CM{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CM{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CM{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 16 - 378"],
                self.get_from_fip(db_path, year, 12483, entcodigo),
            ):
                ws[f"CS{row}"].value = to_excel(
                    datetime.strptime(index, "%Y%m%d"))
                ws[f"CS{row}"].offset(column=1).value = value
                ws[f"CS{row}"].offset(column=2).value = value_fip
                ws[f"CS{row}"].offset(column=1).number_format = "#,##0"
                ws[f"CS{row}"].offset(column=2).number_format = "#,##0"
                ws[f"CS{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 404:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 404"],
                self.get_from_fip(db_path, year, 12244, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 404"],
                self.get_from_fip(db_path, year, 12245, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 404"],
                self.get_from_fip(db_path, year, 12246, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 404"],
                self.get_from_fip(db_path, year, 12249, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 404"],
                self.get_from_fip(db_path, year, 12250, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 6 - 404"],
                self.get_from_fip(db_path, year, 12251, entcodigo),
            ):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 404"],
                self.get_from_fip(db_path, year, 12272, entcodigo),
            ):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 405:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 405"],
                self.get_from_fip(db_path, year, 12258, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 405"],
                self.get_from_fip(db_path, year, 12259, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 405"],
                self.get_from_fip(db_path, year, 12262, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 405"],
                self.get_from_fip(db_path, year, 12263, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 405"],
                self.get_from_fip(db_path, year, 12273, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 406:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 406"],
                self.get_from_fip(db_path, year, 12338, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 406"],
                self.get_from_fip(db_path, year, 12339, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 406"],
                self.get_from_fip(db_path, year, 12341, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 406"],
                self.get_from_fip(db_path, year, 12342, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 407:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 407"],
                self.get_from_fip(db_path, year, 12352, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 407"],
                self.get_from_fip(db_path, year, 12353, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 407"],
                self.get_from_fip(db_path, year, 12379, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 408:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 408"],
                self.get_from_fip(db_path, year, 12064, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 408"],
                self.get_from_fip(db_path, year, 12066, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 408"],
                self.get_from_fip(db_path, year, 12067, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 408"],
                self.get_from_fip(db_path, year, 12069, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 408"],
                self.get_from_fip(db_path, year, 12074, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 6 - 408"],
                self.get_from_fip(db_path, year, 12075, entcodigo),
            ):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 408"],
                self.get_from_fip(db_path, year, 12076, entcodigo),
            ):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 8 - 408"],
                self.get_from_fip(db_path, year, 12059,entcodigo)
                - self.get_from_fip(db_path, year, 12062,entcodigo)):
                ws[f"AW{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AW{row}"].offset(column=1).value = value
                ws[f"AW{row}"].offset(column=2).value = value_fip
                ws[f"AW{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 9 - 408"],
                self.get_from_fip(db_path, year, 12061, entcodigo),
            ):
                ws[f"BC{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"BC{row}"].offset(column=1).value = value
                ws[f"BC{row}"].offset(column=2).value = value_fip
                ws[f"BC{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 10 - 408"],
                self.get_from_fip(db_path, year, 12078, entcodigo),
            ):
                ws[f"BI{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"BI{row}"].offset(column=1).value = value
                ws[f"BI{row}"].offset(column=2).value = value_fip
                ws[f"BI{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BI{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 11 - 408"],
                self.get_from_fip(db_path, year, 12500, entcodigo),
            ):
                ws[f"BO{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"BO{row}"].offset(column=1).value = value
                ws[f"BO{row}"].offset(column=2).value = value_fip
                ws[f"BO{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BO{row}"].offset(column=3).number_format = "#,##0"
                row += 1
        elif qe == 409:
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 1 - 409"],
                self.get_from_fip(db_path, year, 12087, entcodigo),
            ):
                ws[f"G{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"G{row}"].offset(column=1).value = value
                ws[f"G{row}"].offset(column=2).value = value_fip
                ws[f"G{row}"].offset(column=1).number_format = "#,##0"
                ws[f"G{row}"].offset(column=2).number_format = "#,##0"
                ws[f"G{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 2 - 409"],
                self.get_from_fip(db_path, year, 12089, entcodigo),
            ):
                ws[f"M{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"M{row}"].offset(column=1).value = value
                ws[f"M{row}"].offset(column=2).value = value_fip
                ws[f"M{row}"].offset(column=1).number_format = "#,##0"
                ws[f"M{row}"].offset(column=2).number_format = "#,##0"
                ws[f"M{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 3 - 409"],
                self.get_from_fip(db_path, year, 12090, entcodigo),
            ):
                ws[f"S{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"S{row}"].offset(column=1).value = value
                ws[f"S{row}"].offset(column=2).value = value_fip
                ws[f"S{row}"].offset(column=1).number_format = "#,##0"
                ws[f"S{row}"].offset(column=2).number_format = "#,##0"
                ws[f"S{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 4 - 409"],
                self.get_from_fip(db_path, year, 12092, entcodigo),
            ):
                ws[f"Y{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"Y{row}"].offset(column=1).value = value
                ws[f"Y{row}"].offset(column=2).value = value_fip
                ws[f"Y{row}"].offset(column=1).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=2).number_format = "#,##0"
                ws[f"Y{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 5 - 409"],
                self.get_from_fip(db_path, year, 12097, entcodigo),
            ):
                ws[f"AE{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AE{row}"].offset(column=1).value = value
                ws[f"AE{row}"].offset(column=2).value = value_fip
                ws[f"AE{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AE{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 6 - 409"],
                self.get_from_fip(db_path, year, 12098, entcodigo),
            ):
                ws[f"AK{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AK{row}"].offset(column=1).value = value
                ws[f"AK{row}"].offset(column=2).value = value_fip
                ws[f"AK{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AK{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 7 - 409"],
                self.get_from_fip(db_path, year, 12099, entcodigo),
            ):
                ws[f"AQ{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AQ{row}"].offset(column=1).value = value
                ws[f"AQ{row}"].offset(column=2).value = value_fip
                ws[f"AQ{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AQ{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 8 - 409"],
                self.get_from_fip(db_path, year, 12082,entcodigo)
                - self.get_from_fip(db_path, year, 12085,entcodigo)):
                ws[f"AW{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"AW{row}"].offset(column=1).value = value
                ws[f"AW{row}"].offset(column=2).value = value_fip
                ws[f"AW{row}"].offset(column=1).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=2).number_format = "#,##0"
                ws[f"AW{row}"].offset(column=3).number_format = "#,##0"
                row += 1
            row = 14
            for index, value, value_fip in zip(
                df_cruz.index,
                df_cruz["Cruzamento 9 - 409"],
                self.get_from_fip(db_path, year, 12084, entcodigo),
            ):
                ws[f"BC{row}"].value = to_excel(
                    datetime.strptime(index+"01", "%Y%m%d"))
                ws[f"BC{row}"].offset(column=1).value = value
                ws[f"BC{row}"].offset(column=2).value = value_fip
                ws[f"BC{row}"].offset(column=1).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=2).number_format = "#,##0"
                ws[f"BC{row}"].offset(column=3).number_format = "#,##0"
                row += 1

    def save_xl(self, save_path):
        self.ws = self.wb.active
        row = 14
        try:
            for line in self.report:
                self.ws[f"C{row}"].value = line[0]
                self.ws[f"D{row}"].value = line[1]
                row += 1
        except BaseException:
            pass
        save_path = os.path.abspath(
            os.path.join(
                save_path,
                self.ws.title +
                ".xlsx"))
        self.wb.save(save_path)

    def qe_to_csv(self, qe_type, path, file_names):
        if qe_type == "376" or qe_type == 376:
            with open(path + "\\376_Export.csv", "a+") as csv:
                headers = [
                    "ESRSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "TPMOID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESRDATAINICIO",
                    "ESRDATAFIM",
                    "ESRDATAOCORR",
                    "ESRDATAREG",
                    "ESRVALORMOV",
                    "ESRDATACOMUNICA",
                    "ESRCODCESS",
                    "ESRNUMSIN",
                    "ESRVALORMON",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    linha = linha.strip()
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        "/".join([linha[18:20], linha[16:18], linha[12:16]]),
                                        linha[20:23],
                                        linha[23:27],
                                        linha[27:31],
                                        linha[31:35],
                                        "/".join([linha[41:43], linha[39:41], linha[35:39]]),
                                        "/".join([linha[49:51], linha[47:49], linha[43:47]]),
                                        "/".join([linha[57:59], linha[55:57], linha[51:55]]),
                                        "/".join([linha[65:67], linha[63:65], linha[59:63]]),
                                        linha[67:80].replace(",", "."),
                                        "/".join([linha[86:88], linha[84:86], linha[80:84]]),
                                        linha[88:93],
                                        linha[93:113],
                                        linha[113:126].replace(",", "."),
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "377" or qe_type == 377:
            with open(path + "\\377_Export.csv", "a+") as csv:
                headers = [
                    "ESLSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESLDATAINICIO",
                    "ESLDATAFIM",
                    "ESLDATAOCORR",
                    "ESLDATAREG",
                    "ESLVALORMOV",
                    "ESLCODCESS",
                    "ESLNUMSIN",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        "/".join([linha[18:20], linha[16:18], linha[12:16]]),
                                        linha[20:23],
                                        linha[23:27],
                                        linha[27:31],
                                        "/".join([linha[37:39], linha[35:37], linha[31:35]]),
                                        "/".join([linha[45:47], linha[43:45], linha[39:43]]),
                                        "/".join([linha[53:55], linha[51:53], linha[47:51]]),
                                        "/".join([linha[61:63], linha[59:61], linha[55:59]]),
                                        linha[63:76].replace(",", "."),
                                        linha[76:81],
                                        linha[81:101],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "378" or qe_type == 378:
            with open(path + "\\378_Export.csv", "a+") as csv:
                headers = [
                    "ESPSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "TPMOID",
                    "CMPID",
                    "RAMCODIGO",
                    "ESPDATAINICIORO",
                    "ESPDATAFIMRO",
                    "ESPDATAEMISSRO",
                    "ESPVALORMOVRO",
                    "ESPDATAINICIORD",
                    "ESPDATAFIMRD",
                    "ESPDATAEMISSRD",
                    "ESPVALORMOVRD",
                    "ESPCODCESS",
                    "ESPFREQ",
                    "ESPVALORCARO",
                    "ESPVALORCARD",
                    "ESPVALORCIRO",
                    "ESPVALORCIRD",
                    "ESPMOEDA",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        "/".join([linha[18:20], linha[16:18], linha[12:16]]),
                                        linha[20:23],
                                        linha[23:27],
                                        linha[27:31],
                                        linha[31:35],
                                        "/".join([linha[41:43], linha[39:41], linha[35:39]]),
                                        "/".join([linha[49:51], linha[47:49], linha[43:47]]),
                                        "/".join([linha[57:59], linha[55:57], linha[51:55]]),
                                        linha[59:72].replace(",", "."),
                                        "/".join([linha[78:80], linha[76:78], linha[72:76]]),
                                        "/".join([linha[86:88], linha[84:86], linha[80:84]]),
                                        "/".join([linha[94:96], linha[92:94], linha[88:92]]),
                                        linha[96:109].replace(",", "."),
                                        linha[109:114],
                                        linha[114:118],
                                        linha[118:131].replace(",", "."),
                                        linha[131:144].replace(",", "."),
                                        linha[144:157].replace(",", "."),
                                        linha[157:170].replace(",", "."),
                                        linha[170:173],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "404" or qe_type == 404:
            with open(path + "\\404_Export.csv", "a+") as csv:
                headers = [
                    "MSASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MSATIPOPERA",
                    "MSANUMSIN",
                    "MSANUMCONT",
                    "MSATIPOCONT",
                    "MSACODCESS",
                    "MSADATACOMUNICA",
                    "MSADATAREG",
                    "MSADARAOCORR",
                    "MSAVALORMOV",
                    "MSATIPOSIN",
                    "MSAMODCONT",
                    "MSAMOEDA",
                    "MSABASEIND",
                    "MSAVALORMON",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:21],
                                        linha[21:23],
                                        linha[23:24],
                                        linha[24:44],
                                        linha[44:70],
                                        linha[70:71],
                                        linha[71:76],
                                        linha[76:84],
                                        linha[84:92],
                                        linha[92:100],
                                        linha[100:113],
                                        linha[113:114],
                                        linha[114:116],
                                        linha[116:119],
                                        linha[119:120],
                                        linha[120:133],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "405" or qe_type == 405:
            with open(path + "\\405_Export.csv", "a+") as csv:
                headers = [
                    "MSASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MSRNUMSIN",
                    "MSRNUMCONT",
                    "MSRTIPOCONT",
                    "MSRCODCESS",
                    "MSRDATACOMUNICA",
                    "MSRDATAREG",
                    "MSRVALOROCORR",
                    "MSRVALORMOV",
                    "MSRTIPOSIN",
                    "MSRMODCONT",
                    "MSRMOEDA",
                    "MSRDASEIND",
                    "MSRVALORMON",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:21],
                                        linha[21:23],
                                        linha[23:43],
                                        linha[43:69],
                                        linha[69:70],
                                        linha[70:75],
                                        linha[75:83],
                                        linha[83:91],
                                        linha[91:99],
                                        linha[99:112],
                                        linha[112:113],
                                        linha[113:115],
                                        linha[115:118],
                                        linha[118:119],
                                        linha[119:132],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "406" or qe_type == 406:
            with open(path + "\\406_Export.csv", "a+") as csv:
                headers = [
                    "SLASEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "GRACODIGO",
                    "SLATIPOPERA",
                    "SLANUMSIN",
                    "SLANUMCONT",
                    "SLATIPOCONT",
                    "SLACODCESS",
                    "SLADATACOMUNICA",
                    "SLADATAREG",
                    "SLADATAOCORR",
                    "SLAVALORMOVPEN",
                    "SLAVALORMOVTOT",
                    "SLATIPOSIN",
                    "SLAMODCONT",
                    "SLAMOEDA",
                    "SLABASEIND",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:20],
                                        linha[20:21],
                                        linha[21:41],
                                        linha[41:67],
                                        linha[67:68],
                                        linha[68:73],
                                        linha[73:81],
                                        linha[81:89],
                                        linha[89:97],
                                        linha[97:110],
                                        linha[110:123],
                                        linha[123:124],
                                        linha[124:126],
                                        linha[126:129],
                                        linha[129:130],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "407" or qe_type == 407:
            with open(path + "\\407_Export.csv", "a+") as csv:
                headers = [
                    "SLRSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "GRACODIGO",
                    "SLRNUMSIN",
                    "SLRNUMCONT",
                    "SLRTIPOCONT",
                    "SLRCODCESS",
                    "SLRDATACOMUNICA",
                    "SLRDATAREG",
                    "SLRDATAOCORR",
                    "SLRVALORMOVPEN",
                    "SLRVALORMOVTOT",
                    "SLRTIPOSIN",
                    "SLRMODCONT",
                    "SLRMOEDA",
                    "SLRBASEIND",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:20],
                                        linha[20:40],
                                        linha[40:66],
                                        linha[66:67],
                                        linha[67:72],
                                        linha[72:80],
                                        linha[80:88],
                                        linha[88:96],
                                        linha[96:109],
                                        linha[109:122],
                                        linha[122:123],
                                        linha[123:125],
                                        linha[125:128],
                                        linha[128:129],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "408" or qe_type == 408:
            with open(path + "\\408_Export.csv", "a+") as csv:
                headers = [
                    "MPRSEQ",
                    "MRFMESANO",
                    "TPMORESSID",
                    "GRACODIGO",
                    "MPRNUMCONT",
                    "MPRNUMENDOSSO",
                    "MPRCODESS",
                    "MPRTIPOCONT",
                    "MPRMODCONT",
                    "MPRDATACEITE",
                    "MPRDATACONTR",
                    "MPRDATAINICIO",
                    "MPRDATAFIM",
                    "MPRPERCRISCO",
                    "MPRVALORMOV",
                    "MPRVALORMOVCOMIS",
                    "MPRCODCORRET",
                    "MPRVALORMOVCORRET",
                    "MPRVIGMED",
                    "MPRBASEIND",
                    "MPRMOEDA",
                    "MPRTAXACONV",
                    "MPRDATAEMISS ",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:21],
                                        linha[21:23],
                                        linha[23:24],
                                        linha[24:50],
                                        linha[50:56],
                                        linha[56:61],
                                        linha[61:62],
                                        linha[62:64],
                                        linha[64:72],
                                        linha[72:80],
                                        linha[80:88],
                                        linha[88:96],
                                        linha[96:109],
                                        linha[109:115],
                                        linha[115:128],
                                        linha[128:141],
                                        linha[141:146],
                                        linha[146:148],
                                        linha[148:149],
                                        linha[149:152],
                                        linha[152:165],
                                        linha[165:173],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "409" or qe_type == 409:
            with open(path + "\\409_Export.csv", "a+") as csv:
                headers = [
                    "EMFSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "ATVCODIGO",
                    "TPFOPERADOR",
                    "FTRCODIGO",
                    "LCRCODIGO",
                    "TCTCODIGO",
                    "TPECODIGO",
                    "EMFPRAZOFLUXO",
                    "EMFVLREXPRISCO",
                    "EMFCNPJFUNDO",
                    "EMFCODISIN",
                    "EMFCODCUSTODIA",
                    "EMFMULTIPLOFATOR",
                    "EMFTXCONTRATADO",
                    "EMFTXMERCADO",
                    "TPFOPERADORDERIVATIVO",
                    "EMFVLRDERIVATIVO",
                    "EMFCODGRUPO",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:7],
                                        linha[7:12],
                                        linha[12:18],
                                        linha[18:21],
                                        linha[21:23],
                                        linha[23:49],
                                        linha[49:55],
                                        linha[55:60],
                                        linha[60:61],
                                        linha[61:63],
                                        linha[63:71],
                                        linha[71:79],
                                        linha[79:87],
                                        linha[87:95],
                                        linha[95:101],
                                        linha[101:114],
                                        linha[114:127],
                                        linha[127:132],
                                        linha[132:145],
                                        linha[145:147],
                                        linha[147:148],
                                        linha[148:151],
                                        linha[151:164],
                                        linha[164:172],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "419" or qe_type == 419:
            with open(path + "\\419_Export.csv", "a+") as csv:
                headers = [
                    "EMFSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "ATVCODIGO",
                    "TPFOPERADOR",
                    "FTRCODIGO",
                    "LCRCODIGO",
                    "TCTCODIGO",
                    "TPECODIGO",
                    "EMFPRAZOFLUXO",
                    "EMFVLREXPRISCO",
                    "EMFCNPJFUNDO",
                    "EMFCODISIN",
                    "EMFCODCUSTODIA",
                    "EMFMULTIPLOFATOR",
                    "EMFTXCONTRATADO",
                    "EMFTXMERCADO",
                    "TPFOPERADORDERIVATIVO",
                    "EMFVLRDERIVATIVO",
                    "EMFCODGRUPO",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:6],
                                        linha[6:11],
                                        linha[11:19],
                                        linha[19:22],
                                        linha[22:27],
                                        linha[27:28],
                                        linha[28:31],
                                        linha[31:34],
                                        linha[34:36],
                                        linha[36:40],
                                        linha[40:45],
                                        linha[45:60],
                                        linha[60:74],
                                        linha[74:86],
                                        linha[86:98],
                                        linha[98:99],
                                        linha[99:105],
                                        linha[105:111],
                                        linha[111:112],
                                        linha[112:127],
                                        linha[127:133],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "420" or qe_type == 420:
            with open(path + "\\420_Export.csv", "a+") as csv:
                headers = [
                    "EMCSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "DOCCODIGO",
                    "TPFOPERADOR",
                    "FTRCODIGO",
                    "EMCPRAZOFLUXO",
                    "EMCVLREXPRISCO",
                    "EMCMULTIPLOFATOR",
                    "EMCCODGRUPO",
                    "EMCSEMREGISTRO",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:6],
                                        linha[6:11],
                                        linha[11:19],
                                        linha[19:22],
                                        linha[22:27],
                                        linha[27:28],
                                        linha[28:31],
                                        linha[31:36],
                                        linha[36:51],
                                        linha[51:52],
                                        linha[52:58],
                                        linha[58:59],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "421" or qe_type == 421:
            with open(path + "\\421_Export.csv", "a+") as csv:
                headers = [
                    "EMDSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "DODCODIGO",
                    "TPFOPERADOR",
                    "FTRCODIGO",
                    "EMDPRAZOFLUXO",
                    "EMDVLREXPRISCO",
                    "EMDMULTIPLOFATOR",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:6],
                                        linha[6:11],
                                        linha[11:19],
                                        linha[19:22],
                                        linha[22:27],
                                        linha[27:28],
                                        linha[28:31],
                                        linha[31:36],
                                        linha[36:51],
                                        linha[51:52],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "422" or qe_type == 422:
            with open(path + "\\422_Export.csv", "a+") as csv:
                headers = [
                    "EMESEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "EMECODGRUPO",
                    "EMEPEF",
                    "EMEVLRCONTATIVOS",
                    "EMEPERCREVERSAO",
                    "EMEPERCDEDUCAO",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name) as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:6],
                                        linha[6:11],
                                        linha[11:19],
                                        linha[19:22],
                                        linha[22:28],
                                        linha[28:43],
                                        linha[43:58],
                                        linha[58:64],
                                        linha[64:70],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
        if qe_type == "423" or qe_type == 423:
            with open(path + "\\423_Export.csv", "a+") as csv:
                headers = [
                    "EMGSEQ",
                    "ENTCODIGO",
                    "MRFMESANO",
                    "QUAID",
                    "EMGCODGRUPO",
                    "RAMCODIGO",
                    "PLNCODIGO",
                ]
                csv.write(";".join(headers) + "\n")
                for file_name in file_names:
                    try:
                        with open(file_name, encoding="utf-8-sig")  as txt:
                            for linha in txt.readlines():
                                linha = linha.strip()
                                if not (len(linha) == 0 \
                                    or linha == "" \
                                        or linha is None):
                                    structure = [
                                        linha[0:6],
                                        linha[6:11],
                                        linha[11:19],
                                        linha[19:22],
                                        linha[22:28],
                                        linha[28:32],
                                        linha[32:38],
                                    ]
                                    csv.write(";".join(structure) + "\n")
                    except FileNotFoundError:
                        pass
